"""Excel export for report results."""
from __future__ import annotations

import io
from typing import Any

from .execute import ReportResult
from .format import format_display_value


def _display_value(row: dict[str, Any], col) -> str:
    label_key = f"{col.key}__label"
    label_val = row.get(label_key)
    if label_val is not None:
        return str(label_val)
    sym = row.get(f"{col.key}__currency_symbol")
    return format_display_value(
        row.get(col.key),
        fmt=col.format_dict(),
        currency_symbol=sym,
    )


def export_xlsx(result: ReportResult, *, title: str = "Report") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for Excel export (pip install openpyxl)"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = (title[:31] if title else "Report") or "Report"

    headers = [col.label for col in result.columns]
    ws.append(headers)
    for row in result.rows:
        ws.append([_display_value(row, col) for col in result.columns])

    _ALIGN = {"left": "left", "center": "center", "right": "right"}
    for i, col in enumerate(result.columns, start=1):
        letter = get_column_letter(i)
        align = _ALIGN.get(col.format_dict().get("align", "left"), "left")
        cell_align = Alignment(horizontal=align)
        ws[f"{letter}1"].alignment = cell_align
        for r in range(2, ws.max_row + 1):
            ws[f"{letter}{r}"].alignment = cell_align

    for i, _ in enumerate(headers, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = min(40, max(12, len(headers[i - 1]) + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(result: ReportResult) -> str:
    import csv
    import io

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([col.label for col in result.columns])
    for row in result.rows:
        writer.writerow([_display_value(row, col) for col in result.columns])
    return buf.getvalue()
